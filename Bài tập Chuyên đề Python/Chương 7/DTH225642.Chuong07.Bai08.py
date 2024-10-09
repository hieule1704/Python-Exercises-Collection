from openpyxl import load_workbook

def main():
    wb = load_workbook("D:\Python\Chuong7\demo.xlsx")
    print (wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    for row in ws.values:
        for value in row:
            print(value,"\t",end='')
        print("")

if __name__ == "__main__":
    main()